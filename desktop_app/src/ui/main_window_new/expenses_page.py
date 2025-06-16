"""
Expenses Page functionality for the Shiakati Store POS application.
"""

from PyQt5.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, 
    QTableWidget, QTableWidgetItem, QDateEdit, QMessageBox, QDialog,
    QFormLayout, QDialogButtonBox, QDoubleSpinBox, QComboBox,
    QTextEdit, QHeaderView, QWidget, QFileDialog
)
from PyQt5.QtCore import Qt, QDate, QDateTime
from PyQt5.QtGui import QColor

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExpensesPageMixin:
    """Mixin class for the Expenses page functionality."""
    
    def setup_expenses_page(self):
        """Set up the Expenses management page."""
        layout = QVBoxLayout(self.expenses_page)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Header section
        header_layout = QHBoxLayout()
        
        title = QLabel("Expense Tracking")
        title.setStyleSheet("""
            font-size: 20px;
            font-weight: bold;
            color: #2d3436;
        """)
        header_layout.addWidget(title)
        
        header_layout.addStretch()
        
        # Search bar
        self.expense_search = QLineEdit()
        self.expense_search.setPlaceholderText("🔍 Search expenses...")
        self.expense_search.textChanged.connect(self.filter_expenses)
        self.expense_search.setStyleSheet("""
            QLineEdit {
                padding: 8px 15px;
                font-size: 14px;
                min-width: 300px;
                border-radius: 4px;
                border: 1px solid #dcdde1;
            }
        """)
        header_layout.addWidget(self.expense_search)
        
        # Add expense button
        add_btn = QPushButton("➕ Add Expense")
        add_btn.clicked.connect(self.show_add_expense_dialog)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #4834d4;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #686de0;
            }
        """)
        header_layout.addWidget(add_btn)
        
        # Export button
        export_btn = QPushButton("📊 Export")
        export_btn.clicked.connect(self.export_expenses_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        header_layout.addWidget(export_btn)
        
        layout.addLayout(header_layout)
        
        # Date filter section
        filter_layout = QHBoxLayout()
        filter_layout.setContentsMargins(0, 10, 0, 10)
        
        filter_layout.addWidget(QLabel("Filter by date:"))
        
        self.expense_start_date = QDateEdit()
        self.expense_start_date.setDisplayFormat("yyyy-MM-dd")
        self.expense_start_date.setDate(QDate.currentDate().addMonths(-1))
        self.expense_start_date.setCalendarPopup(True)
        filter_layout.addWidget(self.expense_start_date)
        
        filter_layout.addWidget(QLabel("to"))
        
        self.expense_end_date = QDateEdit()
        self.expense_end_date.setDisplayFormat("yyyy-MM-dd")
        self.expense_end_date.setDate(QDate.currentDate())
        self.expense_end_date.setCalendarPopup(True)
        filter_layout.addWidget(self.expense_end_date)
        
        apply_filter_btn = QPushButton("Apply Filter")
        apply_filter_btn.clicked.connect(self.filter_expenses_by_date)
        filter_layout.addWidget(apply_filter_btn)
        
        reset_filter_btn = QPushButton("Reset")
        reset_filter_btn.clicked.connect(self.load_expenses)
        filter_layout.addWidget(reset_filter_btn)
        
        filter_layout.addStretch()
        
        # Summary section
        self.expense_summary = QLabel("Total Expenses: 0 DZD")
        self.expense_summary.setStyleSheet("""
            font-size: 16px;
            font-weight: bold;
            color: #e74c3c;
            margin-left: 10px;
        """)
        filter_layout.addWidget(self.expense_summary)
        
        layout.addLayout(filter_layout)
        
        # Expenses table
        self.expenses_table = QTableWidget()
        self.expenses_table.setColumnCount(7)  # ID, Date, Category, Amount, Description, Payment Method, Actions
        self.expenses_table.setHorizontalHeaderLabels([
            "ID", "Date", "Category", "Amount", "Description", "Payment Method", "Actions"
        ])
        self.expenses_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdde1;
                border-radius: 4px;
                background-color: white;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #f1f2f6;
                min-height: 20px;
            }
            QHeaderView::section {
                background-color: #f8fafc;
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                font-weight: bold;
                color: #475569;
            }
        """)
        
        # Set column widths
        header = self.expenses_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Fixed)  # ID
        header.setSectionResizeMode(1, QHeaderView.Fixed)  # Date
        header.setSectionResizeMode(2, QHeaderView.Fixed)  # Category
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # Amount
        header.setSectionResizeMode(4, QHeaderView.Stretch)  # Description
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # Payment Method
        header.setSectionResizeMode(6, QHeaderView.Fixed)  # Actions
        
        self.expenses_table.setColumnWidth(0, 60)   # ID
        self.expenses_table.setColumnWidth(1, 120)  # Date
        self.expenses_table.setColumnWidth(2, 150)  # Category
        self.expenses_table.setColumnWidth(3, 120)  # Amount
        self.expenses_table.setColumnWidth(5, 150)  # Payment Method
        self.expenses_table.setColumnWidth(6, 150)  # Actions
        
        # Increase row height for better icon visibility
        self.expenses_table.verticalHeader().setDefaultSectionSize(50)
        self.expenses_table.verticalHeader().setVisible(False)  # Hide row numbers
        self.expenses_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.expenses_table.setEditTriggers(QTableWidget.NoEditTriggers)  # Disable editing
        
        layout.addWidget(self.expenses_table)
        
        # Load expenses data
        self.load_expenses()

    def load_expenses_data(self):
        """Wrapper for load_expenses to match the method name in base.py."""
        self.load_expenses()
    
    def load_expenses(self):
        """Load expenses data into the table."""
        try:
            print("\n\n=========== DEBUG: EXPENSES PAGE ===========")
            print("DEBUG: About to call api_client.get_expenses_safely()")
            
            # Check if the method exists before calling
            if not hasattr(self.api_client, 'get_expenses_safely'):
                print("DEBUG: ERROR - get_expenses_safely method doesn't exist in api_client!")
                print(f"DEBUG: Available methods: {[m for m in dir(self.api_client) if not m.startswith('_') and callable(getattr(self.api_client, m))]}")
                
                # Try alternative method names
                if hasattr(self.api_client, 'get_expenses_safe'):
                    print("DEBUG: Found alternative method 'get_expenses_safe', using it instead")
                    expenses = self.api_client.get_expenses_safe()
                elif hasattr(self.api_client, 'get_expenses'):
                    print("DEBUG: Falling back to 'get_expenses' method")
                    expenses = self.api_client.get_expenses()
                else:
                    print("DEBUG: No expenses method found, using empty list")
                    expenses = []
            else:
                # Use the safe method to prevent infinite recursion
                expenses = self.api_client.get_expenses_safely()
            
            print(f"DEBUG: Received expenses data type: {type(expenses)}")
            print(f"DEBUG: Received expenses count: {len(expenses) if expenses else 0}")
            if expenses and len(expenses) > 0:
                print(f"DEBUG: First expense item keys: {list(expenses[0].keys())}")
                print(f"DEBUG: First expense item sample: {expenses[0]}")
                
            self.expenses_table.setRowCount(0)
            
            if not expenses:
                self.expense_summary.setText("Total Expenses: 0 DZD")
                return
                
            total_amount = 0.0
            for expense in expenses:
                print(f"DEBUG: Processing expense: {expense}")
                row = self.expenses_table.rowCount()
                self.expenses_table.insertRow(row)
                
                # Format the expense date
                # Check if we have expense_date (from backend API) or date (local naming)
                date_value = expense.get("expense_date") or expense.get("date")
                if date_value:
                    expense_date = QDateTime.fromString(date_value, Qt.ISODate)
                    formatted_date = expense_date.toString("yyyy-MM-dd")
                else:
                    formatted_date = "Unknown Date"
                
                # Format the amount with color based on amount
                amount = float(expense.get("amount", 0))
                total_amount += amount
                amount_item = QTableWidgetItem(self.format_price(amount))
                amount_item.setForeground(QColor("#e74c3c"))  # Red color for expenses
                
                # Set table items
                self.expenses_table.setItem(row, 0, QTableWidgetItem(str(expense.get("id", ""))))
                self.expenses_table.setItem(row, 1, QTableWidgetItem(formatted_date))
                self.expenses_table.setItem(row, 2, QTableWidgetItem(expense.get("category", "")))
                self.expenses_table.setItem(row, 3, amount_item)
                self.expenses_table.setItem(row, 4, QTableWidgetItem(expense.get("description", "")))
                self.expenses_table.setItem(row, 5, QTableWidgetItem(expense.get("payment_method", "")))
                
                # Create widget for action buttons
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 0, 4, 0)
                actions_layout.setSpacing(8)
                
                # Create edit button
                edit_btn = QPushButton("✏️")
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 3px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                """)
                edit_btn.setFixedSize(30, 25)
                edit_btn.setToolTip("Edit expense")
                edit_btn.clicked.connect(lambda checked, r=row: self.edit_expense(r))
                
                # Create delete button
                delete_btn = QPushButton("🗑️")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 3px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.setFixedSize(30, 25)
                delete_btn.setToolTip("Delete expense")
                delete_btn.clicked.connect(lambda checked, r=row: self.delete_expense(r))
                
                # Add buttons to layout
                actions_layout.addWidget(edit_btn)
                actions_layout.addWidget(delete_btn)
                actions_layout.addStretch()
                
                self.expenses_table.setCellWidget(row, 6, actions_widget)
            
            # Update the total amount display
            self.expense_summary.setText(f"Total Expenses: {self.format_price(total_amount)}")
                
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load expenses: {str(e)}")
            
    def filter_expenses(self):
        """Filter the expenses table based on search text."""
        search_text = self.expense_search.text().lower()
        for row in range(self.expenses_table.rowCount()):
            match = False
            for col in range(self.expenses_table.columnCount() - 1):  # Exclude actions column
                item = self.expenses_table.item(row, col)
                if item and search_text in item.text().lower():
                    match = True
                    break
            self.expenses_table.setRowHidden(row, not match)
            
        # Update the summary with visible expenses only
        total_amount = 0.0
        for row in range(self.expenses_table.rowCount()):
            if not self.expenses_table.isRowHidden(row):
                amount_item = self.expenses_table.item(row, 3)
                if amount_item:
                    total_amount += self.parse_price(amount_item.text())
        
        self.expense_summary.setText(f"Total Expenses: {self.format_price(total_amount)}")

    def filter_expenses_by_date(self):
        """Filter expenses by date range."""
        start_date = self.expense_start_date.date().toString("yyyy-MM-dd")
        end_date = self.expense_end_date.date().addDays(1).toString("yyyy-MM-dd")  # Add one day to include the end date
        
        try:
            filtered_expenses = self.api_client.get_expenses_by_date_range(start_date, end_date)
            if filtered_expenses is None:
                QMessageBox.warning(self, "Error", "Failed to filter expenses by date range")
                return
                
            self.expenses_table.setRowCount(0)  # Clear existing rows
            
            if not filtered_expenses:
                self.expense_summary.setText("Total Expenses: 0 DZD")
                return
                
            # Populate table with filtered data
            total_amount = 0.0
            for expense in filtered_expenses:
                row = self.expenses_table.rowCount()
                self.expenses_table.insertRow(row)
                
                # Format the expense date
                # Check if we have expense_date (from backend API) or date (local naming)
                date_value = expense.get("expense_date") or expense.get("date")
                if date_value:
                    expense_date = QDateTime.fromString(date_value, Qt.ISODate)
                    formatted_date = expense_date.toString("yyyy-MM-dd")
                else:
                    formatted_date = "Unknown Date"
                
                # Format the amount with color based on amount
                amount = float(expense.get("amount", 0))
                total_amount += amount
                amount_item = QTableWidgetItem(self.format_price(amount))
                amount_item.setForeground(QColor("#e74c3c"))  # Red color for expenses
                
                # Set table items
                self.expenses_table.setItem(row, 0, QTableWidgetItem(str(expense.get("id", ""))))
                self.expenses_table.setItem(row, 1, QTableWidgetItem(formatted_date))
                self.expenses_table.setItem(row, 2, QTableWidgetItem(expense.get("category", "")))
                self.expenses_table.setItem(row, 3, amount_item)
                self.expenses_table.setItem(row, 4, QTableWidgetItem(expense.get("description", "")))
                self.expenses_table.setItem(row, 5, QTableWidgetItem(expense.get("payment_method", "")))
                
                # Create widget for action buttons (same as in load_expenses)
                actions_widget = QWidget()
                actions_layout = QHBoxLayout(actions_widget)
                actions_layout.setContentsMargins(4, 0, 4, 0)
                actions_layout.setSpacing(8)
                
                # Create edit button
                edit_btn = QPushButton("✏️")
                edit_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3498db;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 3px;
                    }
                    QPushButton:hover {
                        background-color: #2980b9;
                    }
                """)
                edit_btn.setFixedSize(30, 25)
                edit_btn.setToolTip("Edit expense")
                edit_btn.clicked.connect(lambda checked, r=row: self.edit_expense(r))
                
                # Create delete button
                delete_btn = QPushButton("🗑️")
                delete_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #e74c3c;
                        color: white;
                        border: none;
                        padding: 5px;
                        border-radius: 3px;
                    }
                    QPushButton:hover {
                        background-color: #c0392b;
                    }
                """)
                delete_btn.setFixedSize(30, 25)
                delete_btn.setToolTip("Delete expense")
                delete_btn.clicked.connect(lambda checked, r=row: self.delete_expense(r))
                
                # Add buttons to layout
                actions_layout.addWidget(edit_btn)
                actions_layout.addWidget(delete_btn)
                actions_layout.addStretch()
                
                self.expenses_table.setCellWidget(row, 6, actions_widget)
            
            # Update the total amount display
            self.expense_summary.setText(f"Total Expenses: {self.format_price(total_amount)}")
                
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to filter expenses: {str(e)}")

    def show_add_expense_dialog(self):
        """Show dialog for adding a new expense."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Add New Expense")
        dialog.setMinimumWidth(500)
        layout = QFormLayout(dialog)
        layout.setSpacing(15)
        
        # Create input fields
        category_combo = QComboBox()
        category_combo.addItems([
            "Rent", "Utilities", "Inventory", "Salaries", "Marketing", 
            "Maintenance", "Equipment", "Insurance", "Taxes", "Other"
        ])
        category_combo.setEditable(True)
        
        amount_input = QDoubleSpinBox()
        amount_input.setMaximum(9999999.99)
        amount_input.setDecimals(2)
        amount_input.setSuffix(" DZD")
        amount_input.setValue(0.00)
        
        date_input = QDateEdit()
        date_input.setDisplayFormat("yyyy-MM-dd")
        date_input.setDate(QDate.currentDate())
        date_input.setCalendarPopup(True)
        
        description_input = QTextEdit()
        description_input.setPlaceholderText("Enter expense description")
        description_input.setMaximumHeight(80)
        
        payment_combo = QComboBox()
        payment_combo.addItems(["Cash", "Credit Card", "Bank Transfer", "Check", "Other"])
        payment_combo.setEditable(True)
        
        # Add fields to form
        layout.addRow("Category:", category_combo)
        layout.addRow("Amount:", amount_input)
        layout.addRow("Date:", date_input)
        layout.addRow("Description:", description_input)
        layout.addRow("Payment Method:", payment_combo)
        
        # Add buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
            Qt.Horizontal,
            dialog
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # Get values from form
            category = category_combo.currentText().strip()
            amount = amount_input.value()
            date = date_input.date().toString("yyyy-MM-dd")
            description = description_input.toPlainText().strip()
            payment_method = payment_combo.currentText().strip()
            
            if not category:
                QMessageBox.warning(self, "Error", "Category is required")
                return
                
            if amount <= 0:
                QMessageBox.warning(self, "Error", "Amount must be greater than 0")
                return
            
            # Create expense data
            expense_data = {
                "category": category,
                "amount": amount,
                "expense_date": date,  # Changed from "date" to "expense_date" to match API schema
                "description": description,
                "payment_method": payment_method
            }
            
            try:
                response = self.api_client.create_expense(expense_data)
                if response:
                    QMessageBox.information(self, "Success", "Expense added successfully")
                    self.load_expenses()  # Refresh the expenses table
                else:
                    QMessageBox.warning(self, "Error", "Failed to add expense")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to add expense: {str(e)}")

    def edit_expense(self, row):
        """Edit an expense."""
        expense_id = self.expenses_table.item(row, 0).text()
        current_category = self.expenses_table.item(row, 2).text()
        current_amount = self.parse_price(self.expenses_table.item(row, 3).text())
        current_date_text = self.expenses_table.item(row, 1).text()
        current_date = QDate.fromString(current_date_text, "yyyy-MM-dd")
        current_description = self.expenses_table.item(row, 4).text()
        current_payment_method = self.expenses_table.item(row, 5).text()
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Edit Expense")
        dialog.setMinimumWidth(500)
        layout = QFormLayout(dialog)
        layout.setSpacing(15)
        
        # Create input fields with current values
        category_combo = QComboBox()
        category_combo.addItems([
            "Rent", "Utilities", "Inventory", "Salaries", "Marketing", 
            "Maintenance", "Equipment", "Insurance", "Taxes", "Other"
        ])
        category_combo.setEditable(True)
        category_combo.setCurrentText(current_category)
        
        amount_input = QDoubleSpinBox()
        amount_input.setMaximum(9999999.99)
        amount_input.setDecimals(2)
        amount_input.setSuffix(" DZD")
        amount_input.setValue(current_amount)
        
        date_input = QDateEdit()
        date_input.setDisplayFormat("yyyy-MM-dd")
        date_input.setDate(current_date)
        date_input.setCalendarPopup(True)
        
        description_input = QTextEdit()
        description_input.setText(current_description)
        description_input.setMaximumHeight(80)
        
        payment_combo = QComboBox()
        payment_combo.addItems(["Cash", "Credit Card", "Bank Transfer", "Check", "Other"])
        payment_combo.setEditable(True)
        payment_combo.setCurrentText(current_payment_method)
        
        # Add fields to form
        layout.addRow("Category:", category_combo)
        layout.addRow("Amount:", amount_input)
        layout.addRow("Date:", date_input)
        layout.addRow("Description:", description_input)
        layout.addRow("Payment Method:", payment_combo)
        
        # Add buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.Save | QDialogButtonBox.Cancel,
            Qt.Horizontal,
            dialog
        )
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # Get updated values
            category = category_combo.currentText().strip()
            amount = amount_input.value()
            date = date_input.date().toString("yyyy-MM-dd")
            description = description_input.toPlainText().strip()
            payment_method = payment_combo.currentText().strip()
            
            if not category:
                QMessageBox.warning(self, "Error", "Category is required")
                return
                
            if amount <= 0:
                QMessageBox.warning(self, "Error", "Amount must be greater than 0")
                return
            
            # Update expense data
            expense_data = {
                "category": category,
                "amount": amount,
                "expense_date": date,  # Changed from "date" to "expense_date" to match API schema
                "description": description,
                "payment_method": payment_method
            }
            
            try:
                response = self.api_client.update_expense(expense_id, expense_data)
                if response:
                    QMessageBox.information(self, "Success", "Expense updated successfully")
                    self.load_expenses()  # Refresh the expenses table
                else:
                    QMessageBox.warning(self, "Error", "Failed to update expense")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to update expense: {str(e)}")

    def delete_expense(self, row):
        """Delete an expense."""
        expense_id = self.expenses_table.item(row, 0).text()
        category = self.expenses_table.item(row, 2).text()
        amount = self.expenses_table.item(row, 3).text()
        
        reply = QMessageBox.question(
            self, "Delete Expense",
            f"Are you sure you want to delete this expense?\n\n"
            f"Category: {category}\n"
            f"Amount: {amount}\n",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                response = self.api_client.delete_expense(expense_id)
                if response:
                    QMessageBox.information(self, "Success", "Expense deleted successfully")
                    self.load_expenses()  # Refresh the expenses table
                else:
                    QMessageBox.warning(self, "Error", "Failed to delete expense")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to delete expense: {str(e)}")

    def export_expenses_to_excel(self):
        """Export expenses data to Excel spreadsheet."""
        if self.expenses_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No expenses to export")
            return

        try:
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")  # Red for expenses
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Add headers
            headers = ["ID", "Date", "Category", "Amount", "Description", "Payment Method"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data from visible rows (respecting filters)
            row_num = 2
            total_amount = 0
            for table_row in range(self.expenses_table.rowCount()):
                if not self.expenses_table.isRowHidden(table_row):
                    # Skip the Actions column
                    for col in range(6):  # Only first 6 columns (exclude Actions)
                        item = self.expenses_table.item(table_row, col)
                        if item:
                            cell = ws.cell(row=row_num, column=col+1, value=item.text())
                            cell.border = border
                            
                            # Right align amount column
                            if col == 3:  # Amount column
                                cell.alignment = Alignment(horizontal="right")
                                try:
                                    # Store the numeric amount for total calculation
                                    amount = self.parse_price(item.text())
                                    total_amount += amount
                                except:
                                    pass
                    
                    row_num += 1
            
            # Add total row at the bottom
            row_num += 1
            total_label = ws.cell(row=row_num, column=3, value="Total")
            total_label.font = Font(bold=True)
            total_label.border = border
            
            total_value = ws.cell(row=row_num, column=4, value=self.format_price(total_amount))
            total_value.font = Font(bold=True, color="C0392B")  # Red color
            total_value.border = border
            total_value.alignment = Alignment(horizontal="right")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Ask user for save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if file_path:
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Expenses exported to {file_path}")
                
                # Ask if user wants to open the file
                reply = QMessageBox.question(
                    self, 
                    "Open File", 
                    "Would you like to open the exported file?",
                    QMessageBox.Yes | QMessageBox.No, 
                    QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    self.open_file_with_default_app(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export expenses: {str(e)}")
            import traceback
            traceback.print_exc()

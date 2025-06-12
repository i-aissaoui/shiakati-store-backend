"""
Orders Page functionality for the Shiakati Store POS application.
"""

from PyQt5.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QLineEdit, 
    QTableWidget, QTableWidgetItem, QDateEdit, QMessageBox, QDialog,
    QTextEdit, QHeaderView, QFileDialog, QFormLayout, QComboBox,
    QDialogButtonBox, QApplication
)
from PyQt5.QtCore import Qt, QDate, QDateTime
from PyQt5.QtGui import QColor, QTextDocument
from PyQt5.QtPrintSupport import QPrinter

import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class OrdersPageMixin:
    """Mixin class for the Orders page functionality."""
    
    def setup_orders_page(self):
        """Set up the Orders management page."""
        layout = QVBoxLayout(self.orders_page)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Search and filter section
        search_layout = QHBoxLayout()
        
        self.order_search_input = QLineEdit()
        self.order_search_input.setPlaceholderText("🔍 Search orders...")
        self.order_search_input.textChanged.connect(self.filter_orders)
        self.order_search_input.setStyleSheet("""
            QLineEdit {
                padding: 8px 15px;
                font-size: 14px;
                min-width: 300px;
            }
        """)
        search_layout.addWidget(self.order_search_input)
        
        search_layout.addStretch()
        
        # Date range filters
        date_label = QLabel("Date Range:")
        search_layout.addWidget(date_label)
        
        self.start_date = QDateEdit()
        self.start_date.setDisplayFormat("yyyy-MM-dd")
        self.start_date.setDate(QDate.currentDate().addMonths(-1))  # Default to last month
        self.start_date.setCalendarPopup(True)
        self.start_date.setStyleSheet("""
            QDateEdit {
                padding: 8px;
                font-size: 14px;
                min-width: 120px;
            }
        """)
        search_layout.addWidget(self.start_date)
        
        self.end_date = QDateEdit()
        self.end_date.setDisplayFormat("yyyy-MM-dd")
        self.end_date.setDate(QDate.currentDate())  # Default to today
        self.end_date.setCalendarPopup(True)
        self.end_date.setStyleSheet("""
            QDateEdit {
                padding: 8px;
                font-size: 14px;
                min-width: 120px;
            }
        """)
        search_layout.addWidget(self.end_date)
        
        filter_button = QPushButton("Filter")
        filter_button.clicked.connect(self.filter_orders_by_date)
        search_layout.addWidget(filter_button)
        
        # Export buttons
        export_orders_button = QPushButton("📊 Export Orders")
        export_orders_button.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #218838;
            }
            QPushButton:pressed {
                background-color: #1e7e34;
            }
        """)
        export_orders_button.clicked.connect(self.export_orders_to_excel)
        search_layout.addWidget(export_orders_button)
        
        generate_invoice_button = QPushButton("📋 Generate Invoice")
        generate_invoice_button.setStyleSheet("""
            QPushButton {
                background-color: #007bff;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 15px;
                font-weight: bold;
                font-size: 14px;
                min-width: 120px;
            }
            QPushButton:hover {
                background-color: #0056b3;
            }
            QPushButton:pressed {
                background-color: #004085;
            }
        """)
        generate_invoice_button.clicked.connect(self.generate_monthly_invoice)
        search_layout.addWidget(generate_invoice_button)
        
        layout.addLayout(search_layout)
        
        # Orders table
        self.orders_table = QTableWidget()
        self.orders_table.setColumnCount(10)
        self.orders_table.setHorizontalHeaderLabels([
            "Order ID", "Date", "Customer", "Phone", "Product", "Quantity",
            "Total", "Delivery", "Status", "Notes"
        ])
        # Increase the minimum height for the orders table to take up more window space
        self.orders_table.setMinimumHeight(600)
        self.orders_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdde1;
                border-radius: 4px;
                background-color: white;
                gridline-color: #f1f2f6;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 12px 8px;
                border-bottom: 1px solid #f1f2f6;
                min-height: 20px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
                color: white;
            }
            QHeaderView::section {
                background-color: #f8fafc;
                padding: 12px 8px;
                border: none;
                border-bottom: 2px solid #e2e8f0;
                font-weight: bold;
                color: #475569;
            }
            QTableWidget::item:hover {
                background-color: #f8fafc;
            }
        """)
        self.orders_table.horizontalHeader().setStretchLastSection(True)
        self.orders_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.orders_table.setSelectionMode(QTableWidget.SingleSelection)
        
        # Enable double-click to open order details
        self.orders_table.cellDoubleClicked.connect(self.show_order_details)
        
        # Give the orders table a higher stretch factor to make it take up more space
        layout.addWidget(self.orders_table, 3)  # Higher stretch factor (3) for the orders table
        
        # Order details section
        self.order_details = QTextEdit()
        self.order_details.setReadOnly(True)
        self.order_details.setStyleSheet("""
            QTextEdit {
                border: 1px solid #dcdde1;
                border-radius: 4px;
                background-color: #f8f9fa;
                padding: 10px;
                font-size: 14px;
            }
        """)
        # Add the order details with a lower stretch factor (1) than the orders table (3)
        layout.addWidget(self.order_details, 1)
        
        # Load initial orders data
        self.load_orders_data()

    def load_orders_data(self):
        """Load orders data into the orders table."""
        try:
            orders = self.api_client.get_orders()
            
            self.orders_table.setRowCount(0)  # Clear existing rows
            
            if not orders:
                return
                
            for order in orders:
                try:
                    row = self.orders_table.rowCount()
                    self.orders_table.insertRow(row)
                    
                    # Format the order time
                    order_time = QDateTime.fromString(order.get("order_time"), Qt.ISODateWithMs)
                    formatted_date = order_time.toString("yyyy-MM-dd hh:mm")
                    
                    # Get customer info from the order data
                    customer_name = order.get("customer_name", "N/A")
                    phone_number = order.get("phone_number", "N/A")
                    status = order.get("status", "N/A").capitalize()
                    
                    # Create and style status item
                    status_item = QTableWidgetItem(status)
                    status_color = {
                        "Delivered": "#27ae60",  # Green
                        "Shipped": "#9b59b6",    # Purple
                        "Confirmed": "#2ecc71",  # Light green
                        "Pending": "#f39c12",    # Orange
                        "Cancelled": "#e74c3c"   # Red
                    }.get(status, "#2c3e50")    # Default dark gray
                    status_item.setForeground(QColor(status_color))
                    
                    # Get the first item for display in the main table
                    first_item = order.get("items", [{}])[0] if order.get("items") else {}
                    product_name = first_item.get("product_name", "N/A")
                    size = first_item.get("size", "N/A") 
                    color = first_item.get("color", "N/A")
                    
                    # Ensure N/A is used for None or empty strings
                    if size is None or str(size).strip() == "" or str(size).lower() == "none":
                        size = "N/A"
                    if color is None or str(color).strip() == "" or str(color).lower() == "none":
                        color = "N/A"
                        
                    items_summary = f"{product_name} ({size}, {color})"
                    
                    # Set items with proper formatting
                    self.orders_table.setItem(row, 0, QTableWidgetItem(str(order.get("id", "N/A"))))
                    self.orders_table.setItem(row, 1, QTableWidgetItem(formatted_date))
                    self.orders_table.setItem(row, 2, QTableWidgetItem(customer_name))
                    self.orders_table.setItem(row, 3, QTableWidgetItem(phone_number))
                    self.orders_table.setItem(row, 4, QTableWidgetItem(items_summary))
                    self.orders_table.setItem(row, 5, QTableWidgetItem(str(first_item.get("quantity", 0))))
                    self.orders_table.setItem(row, 6, QTableWidgetItem(self.format_price(float(order.get("total", 0)))))
                    self.orders_table.setItem(row, 7, QTableWidgetItem(str(order.get("delivery_method", "N/A")).capitalize()))
                    self.orders_table.setItem(row, 8, status_item)
                    self.orders_table.setItem(row, 9, QTableWidgetItem(order.get("notes", "")))
                    
                except Exception as item_error:
                    continue
            
            # Adjust column widths
            header = self.orders_table.horizontalHeader()
            # Fixed width columns
            header.setSectionResizeMode(0, QHeaderView.Fixed)  # ID
            header.setSectionResizeMode(1, QHeaderView.Fixed)  # Date
            header.setSectionResizeMode(3, QHeaderView.Fixed)  # Phone
            header.setSectionResizeMode(5, QHeaderView.Fixed)  # Quantity
            header.setSectionResizeMode(6, QHeaderView.Fixed)  # Total 
            header.setSectionResizeMode(7, QHeaderView.Fixed)  # Delivery
            header.setSectionResizeMode(8, QHeaderView.Fixed)  # Status
            
            # Stretching columns
            header.setSectionResizeMode(2, QHeaderView.Stretch)  # Customer name
            header.setSectionResizeMode(4, QHeaderView.Stretch)  # Product
            header.setSectionResizeMode(9, QHeaderView.Stretch)  # Notes
            
            # Set specific widths
            self.orders_table.setColumnWidth(0, 80)   # ID
            self.orders_table.setColumnWidth(1, 150)  # Date
            self.orders_table.setColumnWidth(3, 120)  # Phone
            self.orders_table.setColumnWidth(5, 80)   # Quantity
            self.orders_table.setColumnWidth(6, 150)  # Total - increased significantly for "XXX.XX DZD"
            self.orders_table.setColumnWidth(7, 120)  # Delivery - increased for "Home Delivery"
            self.orders_table.setColumnWidth(8, 100)  # Status
            
            print(f"=== MAIN WINDOW: Successfully loaded {len(orders)} orders ===")
            
        except Exception as e:
            print(f"Error in load_orders_data: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to load orders: {str(e)}")
            import traceback
            traceback.print_exc()  # Print full stack trace for debugging

    def filter_orders(self):
        """Filter the orders table based on search text."""
        search_text = self.order_search_input.text().lower()
        for row in range(self.orders_table.rowCount()):
            match = False
            for col in range(self.orders_table.columnCount()):
                item = self.orders_table.item(row, col)
                if item and search_text in item.text().lower():
                    match = True
                    break
            self.orders_table.setRowHidden(row, not match)

    def filter_orders_by_date(self):
        """Filter orders by the selected date range."""
        start_date = self.start_date.date().toString("yyyy-MM-dd")
        end_date = self.end_date.date().toString("yyyy-MM-dd")
        
        # Add one day to end date to include the end date itself
        end_date_obj = self.end_date.date().addDays(1)
        end_date_inclusive = end_date_obj.toString("yyyy-MM-dd")
        
        try:
            filtered_orders = self.api_client.get_orders_by_date_range(start_date, end_date_inclusive)
            if filtered_orders is None:
                QMessageBox.warning(self, "Error", "Failed to filter orders by date range")
                return
                
            self.orders_table.setRowCount(0)  # Clear existing rows
            
            if not filtered_orders:
                QMessageBox.information(self, "No Results", f"No orders found between {start_date} and {end_date}")
                return
                
            # Same processing as load_orders_data but with filtered orders
            for order in filtered_orders:
                # Process each order the same way as in load_orders_data
                try:
                    row = self.orders_table.rowCount()
                    self.orders_table.insertRow(row)
                    
                    # Format the order time
                    order_time = QDateTime.fromString(order.get("order_time"), Qt.ISODateWithMs)
                    formatted_date = order_time.toString("yyyy-MM-dd hh:mm")
                    
                    # Get customer info from the order data
                    customer_name = order.get("customer_name", "N/A")
                    phone_number = order.get("phone_number", "N/A")
                    status = order.get("status", "N/A").capitalize()
                    
                    # Create and style status item
                    status_item = QTableWidgetItem(status)
                    status_color = {
                        "Delivered": "#27ae60",  # Green
                        "Shipped": "#9b59b6",    # Purple
                        "Confirmed": "#2ecc71",  # Light green
                        "Pending": "#f39c12",    # Orange
                        "Cancelled": "#e74c3c"   # Red
                    }.get(status, "#2c3e50")    # Default dark gray
                    status_item.setForeground(QColor(status_color))
                    
                    # Get the first item for display in the main table
                    first_item = order.get("items", [{}])[0] if order.get("items") else {}
                    product_name = first_item.get("product_name", "N/A")
                    size = first_item.get("size", "N/A") 
                    color = first_item.get("color", "N/A")
                    
                    # Ensure N/A is used for None or empty strings
                    if size is None or str(size).strip() == "" or str(size).lower() == "none":
                        size = "N/A"
                    if color is None or str(color).strip() == "" or str(color).lower() == "none":
                        color = "N/A"
                        
                    items_summary = f"{product_name} ({size}, {color})"
                    
                    # Set items with proper formatting
                    self.orders_table.setItem(row, 0, QTableWidgetItem(str(order.get("id", "N/A"))))
                    self.orders_table.setItem(row, 1, QTableWidgetItem(formatted_date))
                    self.orders_table.setItem(row, 2, QTableWidgetItem(customer_name))
                    self.orders_table.setItem(row, 3, QTableWidgetItem(phone_number))
                    self.orders_table.setItem(row, 4, QTableWidgetItem(items_summary))
                    self.orders_table.setItem(row, 5, QTableWidgetItem(str(first_item.get("quantity", 0))))
                    self.orders_table.setItem(row, 6, QTableWidgetItem(self.format_price(float(order.get("total", 0)))))
                    self.orders_table.setItem(row, 7, QTableWidgetItem(str(order.get("delivery_method", "N/A")).capitalize()))
                    self.orders_table.setItem(row, 8, status_item)
                    self.orders_table.setItem(row, 9, QTableWidgetItem(order.get("notes", "")))
                    
                except Exception as item_error:
                    continue
                    
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to filter orders: {str(e)}")

    def show_order_details(self, row, col):
        """Show details for the selected order."""
        order_id = self.orders_table.item(row, 0).text()
        try:
            order = self.api_client.get_order_details(order_id)
            if not order:
                QMessageBox.warning(self, "Error", "Failed to retrieve order details")
                return
                
            # Format the details as HTML for nice display
            details = f"""
            <h2>Order #{order.get('id')}</h2>
            <p><b>Date:</b> {QDateTime.fromString(order.get('order_time', ''), Qt.ISODateWithMs).toString('yyyy-MM-dd hh:mm')}</p>
            <p><b>Customer:</b> {order.get('customer_name', 'N/A')}</p>
            <p><b>Phone:</b> {order.get('phone_number', 'N/A')}</p>
            <p><b>Status:</b> <span style="color: {self.get_status_color(order.get('status', 'N/A'))};">{order.get('status', 'N/A').capitalize()}</span></p>
            <p><b>Delivery Method:</b> {str(order.get('delivery_method', 'N/A')).capitalize()}</p>
            <p><b>Notes:</b> {order.get('notes', '')}</p>
            
            <h3>Items</h3>
            <table border="1" cellspacing="0" cellpadding="5" style="width:100%; border-collapse: collapse;">
                <tr style="background-color:#f8fafc;">
                    <th>Product</th>
                    <th>Size</th>
                    <th>Color</th>
                    <th>Quantity</th>
                    <th>Price</th>
                    <th>Total</th>
                </tr>
            """
            
            items_total = 0.0
            for item in order.get('items', []):
                price = float(item.get('price', 0))
                quantity = float(item.get('quantity', 0))
                total = price * quantity
                items_total += total
                
                # Ensure empty values are displayed as N/A
                size = item.get('size', 'N/A') 
                if not size or size.lower() == 'none': 
                    size = 'N/A'
                    
                color = item.get('color', 'N/A')
                if not color or color.lower() == 'none': 
                    color = 'N/A'
                
                details += f"""
                <tr>
                    <td>{item.get('product_name', 'N/A')}</td>
                    <td>{size}</td>
                    <td>{color}</td>
                    <td>{quantity}</td>
                    <td>{self.format_price(price)}</td>
                    <td>{self.format_price(total)}</td>
                </tr>
                """
            
            # Add total row
            details += f"""
                <tr style="font-weight: bold; background-color:#f8fafc;">
                    <td colspan="5" align="right">Total:</td>
                    <td>{self.format_price(items_total)}</td>
                </tr>
            </table>
            """
            
            # Display in the details section
            self.order_details.setHtml(details)
            
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to load order details: {str(e)}")

    def get_status_color(self, status):
        """Return color for order status."""
        return {
            "delivered": "#27ae60",  # Green
            "shipped": "#9b59b6",    # Purple
            "confirmed": "#2ecc71",  # Light green
            "pending": "#f39c12",    # Orange
            "cancelled": "#e74c3c"   # Red
        }.get(status.lower(), "#2c3e50")  # Default dark gray
        
    def generate_invoice(self, order_id):
        """Generate invoice for an order."""
        try:
            # Get order details
            order = self.api_client.get_order(order_id)
            if not order:
                QMessageBox.warning(self, "Error", "Order not found")
                return
                
            # Create directory for invoices if it doesn't exist
            invoice_dir = "invoices"
            if not os.path.exists(invoice_dir):
                os.makedirs(invoice_dir)
            
            # Create a unique filename for the invoice
            current_date = datetime.datetime.now().strftime("%Y%m%d")
            file_path = f"{invoice_dir}/Invoice-{order_id}-{current_date}.pdf"
            
            # Set up printer for PDF
            printer = QPrinter()
            printer.setPageSize(QPrinter.A4)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageMargins(15, 15, 15, 15, QPrinter.Unit.Millimeter)
            
            # Create document and set margin
            document = QTextDocument()
            document.setDocumentMargin(10)
            
            # Format the order date
            order_date = order.get("order_time", "")[:19]  # Remove microseconds
            try:
                date_obj = datetime.datetime.fromisoformat(order_date)
                formatted_date = date_obj.strftime("%Y-%m-%d %H:%M")
            except:
                formatted_date = order_date
                
            # Calculate totals
            subtotal = sum(item.get("price", 0) * item.get("quantity", 0) for item in order.get("items", []))
            
            # Build HTML content for the invoice
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    @page {{ 
                        size: A4;
                        margin: 0;
                    }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        margin: 0;
                        padding: 15mm;
                    }}
                    h1, h2 {{ 
                        color: #34495e;
                        margin-bottom: 10px;
                    }}
                    .header {{ 
                        text-align: center;
                        margin-bottom: 20px;
                        border-bottom: 1px solid #ccc;
                        padding-bottom: 10px;
                    }}
                    .company-name {{
                        font-size: 22pt;
                        font-weight: bold;
                        color: #2c3e50;
                    }}
                    .invoice-title {{
                        font-size: 18pt;
                        margin-top: 10px;
                    }}
                    .info-section {{
                        margin-bottom: 20px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-bottom: 20px;
                    }}
                    table, th, td {{
                        border: 1px solid #ddd;
                    }}
                    th {{
                        background-color: #f8f9fa;
                        color: #2c3e50;
                        padding: 10px;
                        text-align: left;
                    }}
                    td {{
                        padding: 8px;
                    }}
                    .total-row {{
                        font-weight: bold;
                    }}
                    .footer {{
                        margin-top: 30px;
                        font-size: 9pt;
                        color: #7f8c8d;
                        text-align: center;
                    }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="company-name">Shiakati Store</div>
                    <div class="invoice-title">INVOICE</div>
                </div>
                
                <div class="info-section" style="display: flex; justify-content: space-between;">
                    <div>
                        <strong>Invoice #:</strong> {order_id}<br>
                        <strong>Date:</strong> {formatted_date}<br>
                        <strong>Status:</strong> {order.get("status", "N/A").upper()}<br>
                    </div>
                    <div>
                        <strong>Customer:</strong> {order.get("customer_name", "N/A")}<br>
                        <strong>Phone:</strong> {order.get("phone_number", "N/A")}<br>
                        <strong>Delivery:</strong> {order.get("delivery_method", "N/A")}<br>
                        <strong>Location:</strong> {order.get("wilaya", "N/A")}, {order.get("commune", "N/A")}
                    </div>
                </div>
                
                <h2>Order Items</h2>
                <table>
                    <tr>
                        <th>Item</th>
                        <th>Size</th>
                        <th>Color</th>
                        <th>Qty</th>
                        <th>Unit Price</th>
                        <th>Total</th>
                    </tr>
            """
            
            # Add items to the invoice
            for item in order.get("items", []):
                product_name = item.get("product_name", "Unknown Product")
                size = item.get("size", "N/A")
                color = item.get("color", "N/A")
                quantity = item.get("quantity", 0)
                unit_price = item.get("price", 0)
                total = quantity * unit_price
                
                html += f"""
                    <tr>
                        <td>{product_name}</td>
                        <td>{size}</td>
                        <td>{color}</td>
                        <td>{quantity}</td>
                        <td>{unit_price:.2f} DZD</td>
                        <td>{total:.2f} DZD</td>
                    </tr>
                """
            
            # Add totals
            html += f"""
                    <tr class="total-row">
                        <td colspan="5" style="text-align: right;"><strong>Subtotal:</strong></td>
                        <td>{subtotal:.2f} DZD</td>
                    </tr>
                    <tr class="total-row">
                        <td colspan="5" style="text-align: right;"><strong>TOTAL:</strong></td>
                        <td>{order.get("total", 0):.2f} DZD</td>
                    </tr>
                </table>
                
                <div>
                    <h2>Additional Information</h2>
                    <p><strong>Notes:</strong> {order.get("notes", "")}</p>
                </div>
                
                <div class="footer">
                    <p>Thank you for your business!</p>
                    <p>Shiakati Store - {datetime.datetime.now().strftime("%Y-%m-%d")}</p>
                </div>
            </body>
            </html>
            """
            
            # Set HTML content and print to PDF
            document.setHtml(html)
            document.print(printer)
            
            # Show success message with option to open the invoice
            reply = QMessageBox.information(
                self,
                "Invoice Generated",
                f"Invoice has been saved to {file_path}",
                QMessageBox.Open | QMessageBox.Ok,
                QMessageBox.Ok
            )
            
            # If user clicked 'Open', open the PDF file
            if reply == QMessageBox.Open:
                import os
                os.system(f"xdg-open {file_path}")
                
            return file_path
                
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to generate invoice: {str(e)}")
            return None
        
    def generate_monthly_invoice(self):
        """Generate a monthly invoice for orders."""
        try:
            # Create a dialog to select the month and year
            dialog = QDialog(self)
            dialog.setWindowTitle("Generate Monthly Invoice")
            dialog.setMinimumWidth(300)
            layout = QVBoxLayout(dialog)
            
            # Add month and year selectors
            form_layout = QFormLayout()
            
            current_date = QDate.currentDate()
            
            month_combo = QComboBox()
            month_names = ["January", "February", "March", "April", "May", "June", 
                          "July", "August", "September", "October", "November", "December"]
            month_combo.addItems(month_names)
            month_combo.setCurrentIndex(current_date.month() - 1)  # Set to current month
            
            year_combo = QComboBox()
            current_year = current_date.year()
            years = [str(year) for year in range(current_year - 5, current_year + 1)]
            year_combo.addItems(years)
            year_combo.setCurrentText(str(current_year))  # Set to current year
            
            form_layout.addRow("Month:", month_combo)
            form_layout.addRow("Year:", year_combo)
            layout.addLayout(form_layout)
            
            # Add buttons
            button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
            button_box.accepted.connect(dialog.accept)
            button_box.rejected.connect(dialog.reject)
            layout.addWidget(button_box)
            
            if dialog.exec_() != QDialog.Accepted:
                return
            
            # Get selected month and year
            month_idx = month_combo.currentIndex() + 1  # 1-12
            year = int(year_combo.currentText())
            
            # Calculate date range for the selected month
            start_date = QDate(year, month_idx, 1)
            end_date = QDate(year, month_idx, 1).addMonths(1).addDays(-1)
            
            # Get orders for the date range
            start_date_str = start_date.toString("yyyy-MM-dd")
            end_date_str = end_date.toString("yyyy-MM-dd")
            
            # Show a progress dialog
            progress = QMessageBox()
            progress.setWindowTitle("Generating Invoice")
            progress.setText("Fetching orders data...")
            progress.setStandardButtons(QMessageBox.NoButton)
            progress.show()
            QApplication.processEvents()
            
            # Fetch orders using the new date range method
            orders = self.api_client.get_orders_by_date_range(start_date_str, end_date_str)
            
            if not orders:
                progress.close()
                QMessageBox.warning(self, "No Orders", f"No orders found for {month_names[month_idx-1]} {year}")
                return
            
            # Create directory for invoices if it doesn't exist
            invoice_dir = "invoices"
            if not os.path.exists(invoice_dir):
                os.makedirs(invoice_dir)
                
            # Create filename for the monthly invoice
            file_path = f"{invoice_dir}/Monthly-Invoice-{year}-{month_idx:02d}.pdf"
            
            # Update progress
            progress.setText("Generating invoice PDF...")
            QApplication.processEvents()
            
            # Set up printer for PDF
            printer = QPrinter()
            printer.setPageSize(QPrinter.A4)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(file_path)
            printer.setPageMargins(15, 15, 15, 15, QPrinter.Unit.Millimeter)
            
            # Create document and set margin
            document = QTextDocument()
            document.setDocumentMargin(10)
            
            # Calculate totals
            total_sales = sum(order.get("total", 0) for order in orders)
            total_items = sum(len(order.get("items", [])) for order in orders)
            
            # Build HTML content for the invoice
            html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <meta charset="UTF-8">
                <style>
                    @page {{ 
                        size: A4;
                        margin: 0;
                    }}
                    body {{ 
                        font-family: Arial, sans-serif;
                        margin: 0;
                        padding: 15mm;
                    }}
                    h1, h2 {{ 
                        color: #34495e;
                    }}
                    .header {{ 
                        text-align: center;
                        margin-bottom: 20px;
                        border-bottom: 1px solid #ccc;
                        padding-bottom: 10px;
                    }}
                    .company-name {{
                        font-size: 22pt;
                        font-weight: bold;
                        color: #2c3e50;
                    }}
                    .invoice-title {{
                        font-size: 18pt;
                        margin-top: 10px;
                    }}
                    .summary-box {{
                        background-color: #f8f9fa;
                        border: 1px solid #e9ecef;
                        border-radius: 4px;
                        padding: 15px;
                        margin-bottom: 20px;
                    }}
                    table {{
                        width: 100%;
                        border-collapse: collapse;
                        margin-bottom: 20px;
                    }}
                    table, th, td {{
                        border: 1px solid #ddd;
                    }}
                    th {{
                        background-color: #f8f9fa;
                        color: #2c3e50;
                        padding: 8px;
                        text-align: left;
                    }}
                    td {{
                        padding: 8px;
                    }}
                    .total-row {{
                        font-weight: bold;
                    }}
                    .footer {{
                        margin-top: 30px;
                        font-size: 9pt;
                        color: #7f8c8d;
                        text-align: center;
                        border-top: 1px solid #ccc;
                        padding-top: 10px;
                    }}
                    .status-pending {{ color: #f39c12; }}
                    .status-completed {{ color: #27ae60; }}
                    .status-canceled {{ color: #e74c3c; }}
                </style>
            </head>
            <body>
                <div class="header">
                    <div class="company-name">Shiakati Store</div>
                    <div class="invoice-title">MONTHLY SALES REPORT</div>
                    <div>{month_names[month_idx-1]} {year}</div>
                </div>
                
                <div class="summary-box">
                    <h2>Monthly Summary</h2>
                    <p><strong>Period:</strong> {start_date.toString("MMMM d, yyyy")} - {end_date.toString("MMMM d, yyyy")}</p>
                    <p><strong>Total Orders:</strong> {len(orders)}</p>
                    <p><strong>Total Items Sold:</strong> {total_items}</p>
                    <p><strong>Total Revenue:</strong> {total_sales:.2f} DZD</p>
                </div>
                
                <h2>Orders</h2>
                <table>
                    <tr>
                        <th>Order ID</th>
                        <th>Date</th>
                        <th>Customer</th>
                        <th>Items</th>
                        <th>Total</th>
                        <th>Status</th>
                    </tr>
            """
            
            # Add orders to the report
            for order in orders:
                order_id = order.get("id", "")
                
                # Format the order date
                order_date = order.get("order_time", "")[:19]  # Remove microseconds
                try:
                    date_obj = datetime.datetime.fromisoformat(order_date)
                    formatted_date = date_obj.strftime("%Y-%m-%d %H:%M")
                except:
                    formatted_date = order_date
                
                customer_name = order.get("customer_name", "N/A")
                item_count = len(order.get("items", []))
                total = order.get("total", 0)
                status = order.get("status", "pending")
                
                # Set status CSS class
                status_class = ""
                if status.lower() == "completed":
                    status_class = "status-completed"
                elif status.lower() == "canceled":
                    status_class = "status-canceled"
                elif status.lower() == "pending":
                    status_class = "status-pending"
                
                html += f"""
                    <tr>
                        <td>{order_id}</td>
                        <td>{formatted_date}</td>
                        <td>{customer_name}</td>
                        <td>{item_count}</td>
                        <td>{total:.2f} DZD</td>
                        <td class="{status_class}">{status.upper()}</td>
                    </tr>
                """
            
            # Complete the HTML
            html += f"""
                </table>
                
                <h2>Status Distribution</h2>
                <table>
                    <tr>
                        <th>Status</th>
                        <th>Count</th>
                        <th>Percentage</th>
                    </tr>
            """
            
            # Calculate status distribution
            status_counts = {}
            for order in orders:
                status = order.get("status", "pending").lower()
                status_counts[status] = status_counts.get(status, 0) + 1
            
            # Add status summary
            for status, count in status_counts.items():
                percentage = (count / len(orders)) * 100 if orders else 0
                
                # Set status CSS class
                status_class = ""
                if status.lower() == "completed":
                    status_class = "status-completed"
                elif status.lower() == "canceled":
                    status_class = "status-canceled"
                elif status.lower() == "pending":
                    status_class = "status-pending"
                
                html += f"""
                    <tr>
                        <td class="{status_class}">{status.upper()}</td>
                        <td>{count}</td>
                        <td>{percentage:.1f}%</td>
                    </tr>
                """
            
            # Complete the HTML
            html += f"""
                </table>
                
                <div class="footer">
                    <p>Report generated on {datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}</p>
                    <p>Shiakati Store - All rights reserved</p>
                </div>
            </body>
            </html>
            """
            
            # Close the progress dialog
            progress.close()
            
            # Set HTML content and print to PDF
            document.setHtml(html)
            document.print(printer)
            
            # Show success message with option to open the invoice
            reply = QMessageBox.information(
                self,
                "Monthly Report Generated",
                f"Monthly sales report has been saved to {file_path}",
                QMessageBox.Open | QMessageBox.Ok,
                QMessageBox.Ok
            )
            
            # If user clicked 'Open', open the PDF file
            if reply == QMessageBox.Open:
                import os
                os.system(f"xdg-open {file_path}")
                
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to generate monthly report: {str(e)}")

    def export_orders_to_excel(self):
        """Export orders data to Excel spreadsheet."""
        if self.orders_table.rowCount() == 0:
            QMessageBox.warning(self, "Warning", "No orders to export")
            return

        try:
            # Create a new workbook and select the active worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Orders"

            # Define styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )
            
            # Add headers
            headers = []
            for col in range(self.orders_table.columnCount()):
                headers.append(self.orders_table.horizontalHeaderItem(col).text())
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Add data from table
            for row in range(self.orders_table.rowCount()):
                for col in range(self.orders_table.columnCount()):
                    item = self.orders_table.item(row, col)
                    if item:
                        cell = ws.cell(row=row+2, column=col+1, value=item.text())
                        cell.border = border
                        if col == 6:  # Total column (price)
                            cell.alignment = Alignment(horizontal="right")
            
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Ask user for save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Save Excel File", "", "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if file_path:
                if not file_path.endswith(".xlsx"):
                    file_path += ".xlsx"
                wb.save(file_path)
                QMessageBox.information(self, "Success", f"Orders exported to {file_path}")
                
                # Ask if user wants to open the file
                reply = QMessageBox.question(
                    self, 
                    "Open File", 
                    "Would you like to open the exported file?",
                    QMessageBox.Yes | QMessageBox.No, 
                    QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    self.open_file_with_default_app(file_path)
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export orders: {str(e)}")
            import traceback
            traceback.print_exc()

    def open_file_with_default_app(self, file_path):
        """Open a file with the default application."""
        try:
            import os
            import platform

            if platform.system() == "Windows":
                os.startfile(file_path)
            elif platform.system() == "Darwin":  # macOS
                os.system(f"open '{file_path}'")
            else:  # Linux and others
                os.system(f"xdg-open '{file_path}'")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to open file: {str(e)}")

    def format_price(self, price):
        """Format price with DZD currency."""
        return f"{price:.2f} DZD"
